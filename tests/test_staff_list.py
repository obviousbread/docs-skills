"""Tests for docs-staff-list and its docs-protocol consumer contract."""

import importlib.util
import subprocess
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "skills" / "docs-staff-list" / "scripts" / "normalize.py"
SPEC = importlib.util.spec_from_file_location("docs_staff_list_normalize", SCRIPT)
normalizer = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(normalizer)
PROTOCOL_SCRIPT = ROOT / "skills" / "docs-protocol" / "generate.py"
PROTOCOL_SPEC = importlib.util.spec_from_file_location("docs_protocol_staff_consumer", PROTOCOL_SCRIPT)
staff_consumer = importlib.util.module_from_spec(PROTOCOL_SPEC)
PROTOCOL_SPEC.loader.exec_module(staff_consumer)


def _source_workbook(path, name="Тестов Тест Тестович"):
    wb = Workbook()
    ws = wb.active
    ws.append(["Сотрудник", "Дата рождения", "Должность", "Подразделение", "Филиал"])
    ws.append([name, "01.02.1980", "специалист", "Отдел А", "Филиал А"])
    wb.save(path)


def test_clean_fio():
    cases = {
        "Тестов Тест Тестович (внеш. совм.)": "Тестов Тест Тестович",
        "Примеров Пример Примерович, Основное место работы, Работа": "Примеров Пример Примерович",
        "Условная Анна АлексеевнаРабота": "Условная Анна Алексеевна",
        "Учебная Наталья Викторовна уборщица": "Учебная Наталья Викторовна",
        "Тестов Тимур Вагиф оглы": "Тестов Тимур Вагиф оглы",
        "Примеров Пример Примерович (метка)": "Примеров Пример Примерович",
        "Чистая Фамилия Отчество": "Чистая Фамилия Отчество",
    }
    for raw, expected in cases.items():
        assert normalizer.clean_fio(raw) == expected


def test_format_date():
    assert normalizer.format_date("21.02.1967") == "21.02.1967"
    assert normalizer.format_date("25.04.82") == "25.04.1982"
    assert normalizer.format_date("30.07.99") == "30.07.1999"
    assert normalizer.format_date("23.10.01") == "23.10.2001"
    assert normalizer.format_date(pd.Timestamp("1972-11-14")) == "14.11.1972"
    assert normalizer.format_date("1972-11-14 00:00:00") == "14.11.1972"
    assert normalizer.format_date("1946", year_only=True) == "1946"
    assert normalizer.format_date(pd.Timestamp("1946-03-08"), year_only=True) == "1946"
    assert normalizer.format_date("") == ""


def test_detect_format():
    hierarchical = pd.DataFrame([
        ["Штатная расстановка"], ["Подразделение"], ["Позиция"], ["Сотрудник, Состояние"],
    ])
    assert normalizer.detect_format(hierarchical)[0] == "shtatnaya"

    table = pd.DataFrame([
        ["№", "Сотрудник", "Должность", "Дата рождения"],
        ["1", "Тестов Тест Тестович", "специалист", "01.01.1980"],
    ])
    assert normalizer.detect_format(table)[0] == "table"

    split = [["1", "Тестова", "Теста", "Тестовна", "23.05.75", "специалист"]] * 5
    assert normalizer.detect_format(pd.DataFrame(split))[0] == "split_fio"


def test_cli_output_is_consumed_by_protocol(tmp_path):
    input_dir = tmp_path / "input"
    older_dir = tmp_path / "staff" / "09.08.2026"
    output_dir = tmp_path / "staff" / "10.08.2026"
    input_dir.mkdir()
    older_dir.mkdir(parents=True)
    output_dir.mkdir(parents=True)
    _source_workbook(input_dir / "branch-a.xlsx")
    _source_workbook(
        older_dir / "Сводный список сотрудников 09.08.2026.xlsx",
        "Устаревший Пример Примерович",
    )
    output = output_dir / "Сводный список сотрудников 10.08.2026.xlsx"

    result = subprocess.run(
        [
            sys.executable, str(SCRIPT),
            "--input-dir", str(input_dir),
            "--title", "Сводный список сотрудников 10.08.2026",
            "--output", str(output),
        ],
        capture_output=True,
        text=True,
        check=False,
    )
    assert result.returncode == 0, result.stderr
    assert "Верификация: OK" in result.stdout

    wb = load_workbook(output, read_only=True, data_only=True)
    ws = wb.active
    assert [cell.value for cell in ws[3]] == [
        "№", "Сотрудник", "Дата рождения", "Должность", "Подразделение", "Филиал",
    ]
    assert ws["B4"].value == "Тестов Тест Тестович"
    wb.close()

    staff = staff_consumer._load_staff(str(tmp_path / "staff"))
    assert staff == [{"lastname": "Тестов", "initials": "Т.Т.", "position": "специалист"}]


def test_protocol_loader_accepts_split_name_columns(tmp_path):
    path = tmp_path / "staff.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Список сотрудников"])
    ws.append([])
    ws.append(["Фамилия", "Имя", "Отчество", "Должность"])
    ws.append(["Примеров", "Пример", "Примерович", "руководитель"])
    wb.save(path)

    assert staff_consumer._load_staff(str(path)) == [
        {"lastname": "Примеров", "initials": "П.П.", "position": "руководитель"},
    ]


def test_cli_requires_output_after_probe(tmp_path):
    _source_workbook(tmp_path / "source.xlsx")
    result = subprocess.run(
        [sys.executable, str(SCRIPT), "--input-dir", str(tmp_path)],
        capture_output=True,
        text=True,
        check=False,
    )
    assert result.returncode == 2
    assert "--output обязателен" in result.stderr

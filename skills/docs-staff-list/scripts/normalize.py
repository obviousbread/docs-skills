#!/usr/bin/env python3
"""
Нормализация и консолидация списков сотрудников из кадровых выгрузок в единый .xlsx.

Главная идея: на входе разнородные .xls/.xlsx от филиалов (разная структура,
разные позиции колонок), на выходе — один сводный файл из 6 колонок:

    № | Сотрудник | Дата рождения | Должность | Подразделение | Филиал

Формат КАЖДОГО файла определяется автоматически по СОДЕРЖИМОМУ (не по имени!).
Три семейства парсеров:

  shtatnaya  — иерархическая «Штатная расстановка» (подразделение→позиция→сотрудник,
               есть метка «Позиция» и колонки Запланировано/Занято/Штатные).
               Варианты: дата в отдельной колонке, только год («Год рождения»),
               год встроен в ФИО («Тестов Т.Т., 1977») без отдельной колонки даты.
  table      — любая ПЛОСКАЯ таблица с заголовком (№/Сотрудник/Должность/Дата рождения).
               Два подрежима, выбираются автоматически:
                 dept-as-column   — есть колонка «Подразделение»/«Отделение» (+может «Филиал»):
                                    каждая строка = сотрудник, подразделение из колонки.
                 dept-as-separator— подразделение идёт строками-разделителями;
                                    строка с числом в №-колонке = сотрудник, иначе = подразделение.
               Покрывает: «Штатные сотрудники» (numbered), плоские выгрузки,
               эталонный 5-столбцовый и УЖЕ СВЕДЁННЫЕ файлы филиалов.
  split_fio  — ФИО разнесено по колонкам Фамилия/Имя/Отчество.

Использование:
    # Консолидация целой папки (формат и филиал — автоопределение):
    python3 normalize.py --input-dir /path --output /path/Сводный.xlsx

    # Консолидация конкретных файлов:
    python3 normalize.py --files "a.xls" "b.xlsx" --input-dir /path -o out.xlsx

    # Переопределения (нужны редко, когда автоопределение ошиблось):
    #   --branch-map "branch-a.xlsx:Филиал А;branch-b.xls:Филиал Б"
    #   --format-map "source.xls:shtatnaya"   --year-only "source.xlsx"
    #   --dedup           — убрать точные дубли (ФИО+Дата рождения), с отчётом
    #   --probe           — только определить форматы/филиалы и выйти (без записи)
"""

import argparse
import glob
import os
import re
import sys
import time
from collections import Counter

try:
    import pandas as pd
except ImportError:
    sys.exit("pandas не установлен. Установите: pip install pandas")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError:
    sys.exit("openpyxl не установлен. Установите: pip install openpyxl")

import warnings
warnings.filterwarnings("ignore")


# ---------------------------------------------------------------------------
# Очистка ФИО
# ---------------------------------------------------------------------------

KNOWN_FIO_PARTICLES = {"оглы", "кызы", "огли", "угли", "де", "фон", "ван"}
fio_changes = []  # (до, после) — для проверки человеком

STATUSES = [
    "Основное место работы, Работа",
    "Основное место работы, Болезнь",
    "Основное место работы, Отсутствие по невыясненным причинам",
    "Внутреннее совместительство, Работа",
    "Внутреннее совместительство, Болезнь",
    "Внешнее совместительство, Работа",
    "Внешнее совместительство, Болезнь",
    "Основное место работы",
    "Внутреннее совместительство",
    "Внешнее совместительство",
    "Работа",
    "Болезнь",
    "Декретный отпуск",
    "Отпуск по уходу за ребенком",
    "Отпуск по беременности и родам",
    "Отпуск основной",
    "Отпуск дополнительный",
    "Отпуск учебный оплачиваемый",
    "Отпуск без сохранения ЗП",
    "Отпуск неоплачиваемый по разрешению работодателя",
    "Отсутствие по невыясненным причинам",
    "Увольнение",
    "Командировка",
]


def clean_fio(fio: str) -> str:
    """Удаляет служебные примечания из поля ФИО, оставляя «Фамилия Имя Отчество»."""
    # 1a: совместительство в скобках
    fio = re.sub(
        r",?\s*\(.*?(совм|совместител|совместительство|внешнее|внутреннее|внешн\.|внутр\.|вн\.).*?\)",
        "", fio, flags=re.IGNORECASE,
    )
    # 1b: совместительство без скобок в конце строки
    fio = re.sub(
        r",?\s+[Вв](?:нешн?|нутр(?:ен(?:н(?:ий|ее|е))?)?|н)\.?\s*(?:совм[а-яё.\-]*|сов(?:\-во|местител[а-яё]*)|св\-во)(?:\s+[\d,]+\s*ст\.?)?.*$",
        "", fio, flags=re.IGNORECASE,
    )
    fio = fio.strip()

    # 2: суффиксы статусов (итеративно, от длинных к коротким)
    changed = True
    while changed:
        changed = False
        for s in STATUSES:
            if fio.endswith(", " + s):
                fio = fio[: -(len(s) + 2)]; changed = True
            elif fio.endswith("," + s):
                fio = fio[: -(len(s) + 1)]; changed = True
    fio = fio.strip().rstrip(",").strip()

    # 2c: статус, приклеенный к ФИО без разделителя («АлексеевнаРабота»)
    for s in STATUSES:
        if fio.endswith(s) and len(fio) > len(s) and fio[-len(s) - 1].islower():
            fio = fio[: -len(s)].strip().rstrip(",").strip()
            break

    # 2b: прочие служебные пометки в скобках (кроме ФИО-частиц)
    def _is_service_paren(m):
        return m.group(0) if m.group(1).strip().lower() in KNOWN_FIO_PARTICLES else ""
    fio = re.sub(r"\s*\(([^)]*)\)", _is_service_paren, fio)

    # 3: обрезать лишние слова после 3 базовых (Фамилия Имя Отчество)
    words = fio.split()
    base = [w for w in words if w.lower() not in KNOWN_FIO_PARTICLES]
    if len(base) > 3:
        kept, n = [], 0
        for w in words:
            if w.lower() in KNOWN_FIO_PARTICLES:
                kept.append(w)
            else:
                n += 1
                if n <= 3:
                    kept.append(w)
                else:
                    break
        fio = " ".join(kept)

    # 4: пробелы
    return re.sub(r"\s+", " ", fio).strip().rstrip(",").strip()


def clean_fio_tracked(fio_raw: str) -> str:
    cleaned = clean_fio(fio_raw)
    if cleaned != fio_raw.strip():
        fio_changes.append((fio_raw.strip(), cleaned))
    return cleaned


def format_date(val, year_only: bool = False):
    """Нормализует дату рождения в 'dd.mm.yyyy' (или год, если year_only)."""
    if pd.isna(val) or str(val).strip() in ("", "nan"):
        return ""
    if hasattr(val, "strftime"):
        try:
            return str(val.year) if year_only else val.strftime("%d.%m.%Y")
        except Exception:
            pass
    s = str(val).strip()
    if year_only:
        m = re.search(r"(\d{4})", s)
        return m.group(1) if m else s
    if re.match(r"\d{4}-\d{2}-\d{2}", s):
        try:
            return pd.to_datetime(s).strftime("%d.%m.%Y")
        except Exception:
            pass
    m = re.match(r"^(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{2})$", s)  # dd.mm.yy
    if m:
        d, mo, yy = m.groups(); yy = int(yy)
        # ponytail: 2-значный год сотрудника — порог 40 (>=40→19xx, иначе 20xx).
        year = 1900 + yy if yy >= 40 else 2000 + yy
        return f"{int(d):02d}.{int(mo):02d}.{year}"
    m = re.match(r"^(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{4})$", s)  # dd.mm.yyyy
    if m:
        d, mo, y = m.groups()
        return f"{int(d):02d}.{int(mo):02d}.{y}"
    return s


def extract_position(text: str) -> str:
    """Должность из строки, отрезая путь /Подразделение/ и хвост после запятой."""
    text = str(text).strip()
    text = re.sub(r"\s*/[^/]+(/[^/]+)*\s*/?\s*$", "", text).strip()
    return text.split(",")[0].strip()


def _is_int(v) -> bool:
    try:
        int(float(str(v).strip()))
        return True
    except (ValueError, TypeError):
        return False


# ---------------------------------------------------------------------------
# Чтение файла и автоопределение формата
# ---------------------------------------------------------------------------

def load_df(filepath: str):
    return pd.read_excel(filepath, header=None)


def _cell(df, i, c):
    if c is None or i >= len(df) or c >= len(df.columns):
        return ""
    v = df.iat[i, c]
    if pd.isna(v):
        return ""
    s = str(v).strip()
    return "" if s == "nan" else s


def _head_texts(df, n=14):
    out = []
    for i in range(min(n, len(df))):
        for c in range(min(len(df.columns), 16)):
            t = _cell(df, i, c).lower()
            if t:
                out.append(t)
    return out


def _looks_like_split_fio(df) -> bool:
    """Данные вида: №(int) | Фамилия | Имя | Отчество | дата | должность (без заголовка)."""
    checked = hits = 0
    for i in range(min(45, len(df))):
        if not _is_int(_cell(df, i, 0)):
            continue
        checked += 1
        a, b, c4 = _cell(df, i, 1), _cell(df, i, 2), _cell(df, i, 4)
        if a and b and " " not in a and " " not in b and \
           re.search(r"\d{1,2}[.\-/]\d{1,2}[.\-/]\d{2,4}", c4):
            hits += 1
        if checked >= 10:
            break
    return checked > 0 and hits >= max(2, checked // 2)


def detect_format(df):
    """Возвращает (format_key, flags). format_key ∈ {shtatnaya, table, split_fio}."""
    head = _head_texts(df, 14)
    flags = {}
    plan_kw = any(t in ("запланировано", "занято", "зянято", "свободно",
                        "штатные", "штаные") for t in head)
    # «Позиция» как самостоятельная метка + «Сотрудник, Состояние/Вид занятости» —
    # уникальная подпись иерархической «Штатной расстановки» (в плоских таблицах
    # должность называется «Должность», а не «Позиция»).
    has_pozic = any(t == "позиция" for t in head)
    has_sost = any(t.startswith("сотрудник,") for t in head)
    has_title = any("штатная расстановка" in t for t in head)

    if has_pozic or has_sost or (has_title and plan_kw):
        if any("год рожд" in t for t in head):
            flags["year_only"] = True
        return "shtatnaya", flags

    if (any(t == "фамилия" for t in head) and any(t == "имя" for t in head)) \
       or _looks_like_split_fio(df):
        return "split_fio", flags

    return "table", flags


def detect_table_columns(df):
    """Находит строку заголовка плоской таблицы и сопоставляет колонки.

    Возвращает (header_idx, colmap) где colmap может содержать ключи
    num, fio, pos, dob, dept, branch. header_idx=None если заголовок не найден.
    """
    header = None
    for i in range(min(18, len(df))):
        vals = [_cell(df, i, c).lower() for c in range(len(df.columns))]
        has_fio = any(v == "сотрудник" or v == "фио" or v.startswith("сотрудник") or "ф.и.о" in v
                      for v in vals if v)
        has_other = any("должн" in v or "позиц" in v or "рожд" in v or
                        "подразд" in v or "отделение" in v for v in vals if v)
        if has_fio and has_other:
            header = i
            break
    if header is None:
        return None, {}

    cm = {}
    for c in range(len(df.columns)):
        v = _cell(df, header, c).lower()
        if not v:
            continue
        if v in ("№", "n", "nn", "no") or "п/п" in v:
            cm.setdefault("num", c)
        elif v == "сотрудник" or v == "фио" or v.startswith("сотрудник") or "ф.и.о" in v:
            cm.setdefault("fio", c)
        elif "должн" in v or v == "позиция":
            cm.setdefault("pos", c)
        elif "рожд" in v:
            cm.setdefault("dob", c)
            if "год рожд" in v:
                cm["dob_year_only"] = True
        elif "подразд" in v or "отделение" in v:
            cm.setdefault("dept", c)
        elif "филиал" in v:
            cm.setdefault("branch", c)
    return header, cm


# ---------------------------------------------------------------------------
# Парсеры
# ---------------------------------------------------------------------------

def parse_table(df):
    """Плоская таблица. Авто: dept-as-column либо dept-as-separator.

    Возвращает (rows, meta). meta содержит has_branch (bool), skipped_continuation (int).
    """
    header, cm = detect_table_columns(df)
    if header is None or "fio" not in cm:
        # Заголовок не распознан — позиционный фолбэк (№,ФИО,дата,должность,подразделение).
        header, cm = -1, {"num": 0, "fio": 1, "dob": 2, "pos": 3, "dept": 4}

    num_c, fio_c = cm.get("num", 0), cm["fio"]
    pos_c, dob_c = cm.get("pos"), cm.get("dob")
    dept_c, branch_c = cm.get("dept"), cm.get("branch")
    year_only = cm.get("dob_year_only", False)
    dept_as_col = dept_c is not None

    rows, current_dept, skipped = [], "", 0
    for i in range(header + 1, len(df)):
        fio_raw = _cell(df, i, fio_c)
        if dept_as_col:
            if not fio_raw:
                # строка-продолжение (доп. должность того же человека) или пустая
                if _cell(df, i, pos_c):
                    skipped += 1
                continue
            d = _cell(df, i, dept_c)
            if d:
                current_dept = d
            rows.append({
                "Сотрудник": clean_fio_tracked(fio_raw),
                "Дата рождения": format_date(df.iat[i, dob_c] if dob_c is not None and dob_c < len(df.columns) else "", year_only),
                "Должность": _cell(df, i, pos_c),
                "Подразделение": current_dept,
                "Филиал": _cell(df, i, branch_c) if branch_c is not None else "",
            })
        else:
            num_val = _cell(df, i, num_c)
            if _is_int(num_val) and fio_raw:
                rows.append({
                    "Сотрудник": clean_fio_tracked(fio_raw),
                    "Дата рождения": format_date(df.iat[i, dob_c] if dob_c is not None and dob_c < len(df.columns) else "", year_only),
                    "Должность": _cell(df, i, pos_c),
                    "Подразделение": current_dept,
                    "Филиал": "",
                })
            else:
                c0 = _cell(df, i, 0)
                # строка-разделитель подразделения (текст, не число, не итог)
                if c0 and not _is_int(c0) and not c0.lower().startswith("всего"):
                    current_dept = re.sub(r"\s+", " ", c0).strip()
    return rows, {"has_branch": branch_c is not None, "skipped_continuation": skipped}


def parse_split_fio(df):
    """ФИО разнесено по Фамилия/Имя/Отчество. Подразделения — строки-разделители."""
    rows, current_dept = [], ""
    for i in range(len(df)):
        c0 = _cell(df, i, 0)
        if not c0:
            continue
        if _is_int(c0):
            full = " ".join(filter(None, [_cell(df, i, 1), _cell(df, i, 2), _cell(df, i, 3)]))
            rows.append({
                "Сотрудник": clean_fio_tracked(full),
                "Дата рождения": format_date(df.iat[i, 4] if 4 < len(df.columns) else ""),
                "Должность": _cell(df, i, 5),
                "Подразделение": current_dept,
                "Филиал": "",
            })
        elif len(c0) > 2 and c0.lower() not in ("фамилия",):
            current_dept = c0
    return rows, {"has_branch": False, "skipped_continuation": 0}


def parse_shtatnaya(df, date_col=None, year_only=False):
    """Иерархическая «Штатная расстановка»."""
    # автодетект колонки даты рождения
    if date_col is None:
        for idx in range(min(9, len(df))):
            for c in range(len(df.columns)):
                val = _cell(df, idx, c).lower()
                if "рожд" in val:
                    date_col = c
                    if "год рожд" in val:
                        year_only = True
                    break
            if date_col is not None:
                break
        if date_col is None:
            embedded = any(re.search(r",\s*\d{4}\s*$", _cell(df, idx, 0))
                           for idx in range(9, min(30, len(df))))
            if not embedded:
                date_col = 4  # фолбэк

    # plan_col (Запланировано/Штатные)
    plan_col = None
    for idx in range(min(9, len(df))):
        for c in range(len(df.columns)):
            val = _cell(df, idx, c).lower()
            if "запланировано" in val or "штатные" in val or "штаные" in val:
                plan_col = c
                break
        if plan_col is not None:
            break
    if plan_col is None:
        plan_col = 5

    # --- режим «год встроен в ФИО» (date_col не найдена) ---
    if date_col is None:
        status_re = re.compile(r",\s*(?:Работа|Болезнь|Отпуск|Увольнение|Декрет|Отсутствие).*$",
                               re.IGNORECASE)
        rows, current_dept, current_pos = [], "", ""
        for i in range(9, len(df)):
            col0 = _cell(df, i, 0)
            if not col0:
                continue
            has_plan = bool(_cell(df, i, plan_col))
            if has_plan:
                is_pos = False
                for look in range(i + 1, min(i + 3, len(df))):
                    if not _cell(df, look, 0):
                        continue
                    is_pos = not bool(_cell(df, look, plan_col))
                    break
                if is_pos:
                    current_pos = col0.split(",")[0].strip()
                else:
                    current_dept = re.sub(r"\s+", " ", re.sub(r'\s*"[^"]*"', "", col0)).strip()
                continue
            if status_re.search(col0):
                continue
            m = re.search(r",\s*(\d{4})\s*$", col0)
            dob = ""
            if m:
                dob = m.group(1)
                col0 = col0[: m.start()].strip()
            name = clean_fio_tracked(col0)
            if name:
                rows.append({"Сотрудник": name, "Дата рождения": dob,
                             "Должность": current_pos, "Подразделение": current_dept,
                             "Филиал": ""})
        return rows, {"has_branch": False, "skipped_continuation": 0}

    # --- стандартный режим (отдельная колонка даты) ---
    employee_rows = {i for i in range(9, len(df)) if _cell(df, i, date_col)}
    rows, current_dept, current_pos = [], "", ""

    for i in range(9, len(df)):
        col0 = _cell(df, i, 0)
        if not col0:
            continue
        if i in employee_rows:
            dval = df.iat[i, date_col]
            rows.append({
                "Сотрудник": clean_fio_tracked(col0),
                "Дата рождения": format_date(dval, year_only),
                "Должность": current_pos,
                "Подразделение": current_dept,
                "Филиал": "",
            })
        elif "/" in col0:
            current_pos = extract_position(col0)
        else:
            is_pos = False
            for look in range(i + 1, min(i + 5, len(df))):
                if look in employee_rows:
                    is_pos = True
                    break
                if _cell(df, look, 0):
                    break
            if is_pos:
                current_pos = extract_position(col0)
            elif _cell(df, i, plan_col):
                current_dept = col0

    # починка пустых должностей обратным поиском
    for emp in rows:
        if emp["Должность"]:
            continue
        surname = emp["Сотрудник"].split()[0] if emp["Сотрудник"] else ""
        if not surname:
            continue
        for i in range(9, len(df)):
            if i in employee_rows and surname in _cell(df, i, 0):
                for j in range(i - 1, max(i - 25, 8), -1):
                    prev = _cell(df, j, 0)
                    if not prev or prev == "," or j in employee_rows:
                        continue
                    if "/" in prev:
                        emp["Должность"] = extract_position(prev); break
                    if _cell(df, j, plan_col) and prev != emp["Подразделение"]:
                        emp["Должность"] = extract_position(prev); break
                break
    return rows, {"has_branch": False, "skipped_continuation": 0}


PARSERS = {"shtatnaya": parse_shtatnaya, "table": parse_table, "split_fio": parse_split_fio}


def parse_file(filepath, fmt=None, year_only=False, date_col=None):
    """Читает файл, определяет формат (если не задан) и парсит. Возвращает (rows, fmt, meta)."""
    df = load_df(filepath)
    flags = {}
    if fmt is None:
        fmt, flags = detect_format(df)
    if fmt == "shtatnaya":
        rows, meta = parse_shtatnaya(df, date_col=date_col,
                                     year_only=year_only or flags.get("year_only", False))
    elif fmt == "split_fio":
        rows, meta = parse_split_fio(df)
    else:
        rows, meta = parse_table(df)
    return rows, fmt, meta


# ---------------------------------------------------------------------------
# Запись .xlsx
# ---------------------------------------------------------------------------

THIN = Border(*([Side(style="thin")] * 4))


def write_consolidated_xlsx(out_path, title, all_rows):
    wb = Workbook(); ws = wb.active; ws.title = "Сотрудники"
    hf = Font(name="Arial", bold=True, size=11)
    fill = PatternFill("solid", fgColor="D9E1F2")
    hc = Alignment(horizontal="center", vertical="center", wrap_text=True)
    df_ = Font(name="Arial", size=10)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A1:F1")
    ws["A1"] = title
    ws["A1"].font = Font(name="Arial", bold=True, size=13)
    ws["A1"].alignment = center

    for ci, h in enumerate(["№", "Сотрудник", "Дата рождения", "Должность",
                            "Подразделение", "Филиал"], 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = hf, fill, hc, THIN

    for ri, emp in enumerate(all_rows):
        r = 4 + ri
        ws.cell(row=r, column=1, value=ri + 1)
        ws.cell(row=r, column=2, value=emp["Сотрудник"])
        ws.cell(row=r, column=3, value=emp["Дата рождения"])
        ws.cell(row=r, column=4, value=emp["Должность"])
        ws.cell(row=r, column=5, value=emp["Подразделение"])
        ws.cell(row=r, column=6, value=emp.get("Филиал", ""))
        for ci in range(1, 7):
            cell = ws.cell(row=r, column=ci)
            cell.font, cell.border = df_, THIN
            cell.alignment = center if ci in (1, 3) else left

    for col, w in {"A": 6, "B": 38, "C": 16, "D": 45, "E": 50, "F": 30}.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:F{3 + len(all_rows)}"
    wb.save(out_path)
    return len(all_rows)


# ---------------------------------------------------------------------------
# Верификация
# ---------------------------------------------------------------------------

def verify_rows(rows, consolidate=False):
    issues = []
    for i, emp in enumerate(rows):
        fio = emp.get("Сотрудник", "")
        if not fio.strip():
            issues.append(f"  Строка {i + 4}: пустое ФИО")
        if re.search(r"Работа|Болезнь|Отпуск|совм", fio, re.IGNORECASE):
            issues.append(f"  Строка {i + 4}: остаточный статус в ФИО: [{fio}]")
        base = [w for w in fio.split() if w.lower() not in KNOWN_FIO_PARTICLES]
        if len(base) > 3:
            issues.append(f"  Строка {i + 4}: >3 слов в ФИО: [{fio}]")
        if not emp.get("Должность", "").strip():
            issues.append(f"  Строка {i + 4}: пустая должность: {fio}")
        if not emp.get("Подразделение", "").strip():
            issues.append(f"  Строка {i + 4}: пустое подразделение: {fio}")
        if consolidate and not emp.get("Филиал", "").strip():
            issues.append(f"  Строка {i + 4}: пустой филиал: {fio}")
    return issues


def find_duplicates(rows):
    key = lambda e: (e["Сотрудник"].strip().lower(), str(e.get("Дата рождения", "")).strip())
    c = Counter(key(e) for e in rows if e["Сотрудник"].strip())
    return {k: v for k, v in c.items() if v > 1}


def print_fio_changes():
    if not fio_changes:
        print("\nИзменений ФИО: 0")
        return
    print(f"\nИЗМЕНЁННЫЕ ФИО ({len(fio_changes)}):")
    for a, b in fio_changes:
        print(f"  {a}  ->  {b}")
    fio_changes.clear()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_kv(text):
    out = {}
    for e in (text or "").split(";"):
        e = e.strip()
        if ":" in e:
            k, v = e.split(":", 1)
            out[k.strip()] = v.strip()
    return out


def branch_from_filename(name):
    """Грубая догадка о филиале из имени файла (для проверки человеком)."""
    base = os.path.splitext(name)[0]
    base = re.sub(r"\d", "", base)
    base = re.sub(r"(?i)\b(сводный|список|сотрудник\w*|штат\w*|таблица|для|ИА|осн|внеш|внутр|на|г|год)\b", "", base)
    base = re.sub(r"[._+]+", " ", base)
    return re.sub(r"\s+", " ", base).strip(" .-") or base


def gather_files(input_dir, files):
    if files:
        return files
    found = []
    for ext in ("*.xls", "*.xlsx", "*.XLS", "*.XLSX"):
        found += [os.path.basename(p) for p in glob.glob(os.path.join(input_dir, ext))]
    # исключаем уже готовые сводные файлы в той же папке
    found = [f for f in found if not f.lower().startswith("сводный")]
    return sorted(set(found))


def main():
    ap = argparse.ArgumentParser(description="Нормализация/консолидация списков сотрудников")
    ap.add_argument("--input-dir", required=True)
    ap.add_argument("--files", nargs="*", default=[], help="Конкретные файлы (иначе вся папка)")
    ap.add_argument("--output", "-o", default=None, help="Путь к сводному .xlsx")
    ap.add_argument("--title", default="Сводный список сотрудников")
    ap.add_argument("--branch-map", default="", help="'файл:Филиал;...' (переопределение)")
    ap.add_argument("--format-map", default="", help="'файл:формат;...' (переопределение автодетекта)")
    ap.add_argument("--year-only", default="", help="Файлы с годом вместо даты, через ;")
    ap.add_argument("--date-col", default="", help="'файл:номер' (0-based), переопределение")
    ap.add_argument("--dedup", action="store_true", help="Убрать точные дубли (ФИО+ДР)")
    ap.add_argument("--probe", action="store_true", help="Только определить форматы/филиалы")
    args = ap.parse_args()

    files = gather_files(args.input_dir, args.files)
    if not files:
        sys.exit("Нет входных файлов.")
    branch_map = parse_kv(args.branch_map)
    format_map = parse_kv(args.format_map)
    date_col_map = parse_kv(args.date_col)
    year_only = {f.strip() for f in args.year_only.split(";") if f.strip()}

    t0 = time.time()
    all_rows, per_file = [], []
    for fn in files:
        fp = os.path.join(args.input_dir, fn)
        if not os.path.exists(fp):
            print(f"  ПРОПУЩЕН (нет файла): {fn}")
            continue
        ft = time.time()
        rows, fmt, meta = parse_file(
            fp,
            fmt=format_map.get(fn),
            year_only=fn in year_only,
            date_col=int(date_col_map[fn]) if fn in date_col_map else None,
        )
        # филиал: переопределение → колонка Филиал в файле → догадка по имени
        if fn in branch_map:
            for e in rows:
                e["Филиал"] = branch_map[fn]
            br = branch_map[fn]
        elif meta["has_branch"] and all(e.get("Филиал", "").strip() for e in rows):
            br = "(из колонки Филиал)"
        else:
            guess = branch_from_filename(fn)
            for e in rows:
                if not e.get("Филиал", "").strip():
                    e["Филиал"] = guess
            br = guess + " (?из имени файла)"
        all_rows.extend(rows)
        per_file.append((fn, fmt, len(rows), br, meta["skipped_continuation"]))
        print(f"  {fn[:48]:<48} fmt={fmt:<10} строк={len(rows):<5} {round(time.time()-ft,2)}s  → {br}")

    if args.probe:
        print(f"\nProbe готов за {round(time.time()-t0,2)}s. Записи не было.")
        return

    if not args.output:
        ap.error("--output обязателен при консолидации")

    if args.dedup:
        seen, dedup = set(), []
        for e in all_rows:
            k = (e["Сотрудник"].strip().lower(), str(e.get("Дата рождения", "")).strip())
            if k in seen:
                continue
            seen.add(k); dedup.append(e)
        print(f"\nDedup: {len(all_rows)} → {len(dedup)} (убрано {len(all_rows)-len(dedup)})")
        all_rows = dedup

    dups = find_duplicates(all_rows)
    issues = verify_rows(all_rows, consolidate=True)
    out = args.output
    write_consolidated_xlsx(out, args.title, all_rows)
    elapsed = time.time() - t0

    print_fio_changes()
    print(f"\n{'='*64}\nИТОГ за {round(elapsed,2)}s")
    print(f"  Файлов: {len(per_file)}   Сотрудников: {len(all_rows)}")
    print(f"  Сохранён: {out}")
    skipped_total = sum(p[4] for p in per_file)
    if skipped_total:
        print(f"  Пропущено строк-продолжений (доп.должности): {skipped_total}")
    print("\n  По филиалам:")
    for br, cnt in Counter(e.get("Филиал", "") for e in all_rows).most_common():
        print(f"    {cnt:>5}  {br}")
    if dups:
        print(f"\n  ДУБЛИ (ФИО+ДР), {len(dups)} шт. (--dedup чтобы убрать):")
        for (name, dob), n in list(dups.items())[:15]:
            print(f"    {n}×  {name} {dob}")
        if len(dups) > 15:
            print(f"    ... ещё {len(dups)-15}")
    if issues:
        print(f"\n  ПРОБЛЕМЫ ({len(issues)}):")
        for s in issues[:25]:
            print(s)
        if len(issues) > 25:
            print(f"    ... ещё {len(issues)-25}")
    else:
        print("\n  Верификация: OK")
    # перф-метрика
    if elapsed > 30:
        print(f"\n  ⚠ Медленно ({round(elapsed,1)}s > 30s) — проверьте размер файлов.")


if __name__ == "__main__":
    main()
